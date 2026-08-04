from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count
from django.http import HttpResponse
from apps.bookings.models import Booking, BookingItem
from apps.tours.models import Tour
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

@staff_member_required
def admin_analytics_view(request):
    # Calculations
    total_sales = Booking.objects.filter(status='confirmed').aggregate(Sum('grand_total'))['grand_total__sum'] or 0.0
    bookings_count = Booking.objects.count()
    confirmed_count = Booking.objects.filter(status='confirmed').count()
    cancelled_count = Booking.objects.filter(status='cancelled').count()
    conversion_rate = (confirmed_count / bookings_count * 100) if bookings_count > 0 else 0.0

    # Recent Bookings
    recent_bookings = Booking.objects.all().order_by('-created_at')[:8]

    # Chart Data 1: Popular Tours
    popular_items = BookingItem.objects.filter(item_type='tour_package', booking__status='confirmed') \
        .values('tour_package__tour__name') \
        .annotate(total_sold=Sum('quantity')) \
        .order_by('-total_sold')[:5]
    
    chart_tour_labels = [item['tour_package__tour__name'] for item in popular_items]
    chart_tour_data = [item['total_sold'] for item in popular_items]

    # Chart Data 2: Monthly Revenue (Last 6 Months)
    today = datetime.today()
    monthly_sales = []
    monthly_labels = []
    
    for i in range(5, -1, -1):
        date_offset = today - timedelta(days=i*30)
        month_start = date_offset.replace(day=1, hour=0, minute=0, second=0)
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year+1, month=1) - timedelta(seconds=1)
        else:
            month_end = month_start.replace(month=month_start.month+1) - timedelta(seconds=1)

        rev = Booking.objects.filter(status='confirmed', created_at__range=(month_start, month_end)) \
            .aggregate(Sum('grand_total'))['grand_total__sum'] or 0.0
        
        monthly_sales.append(float(rev))
        monthly_labels.append(month_start.strftime('%B %Y'))

    context = {
        "title": "Girasol Analytics Dashboard",
        "total_sales": total_sales,
        "bookings_count": bookings_count,
        "confirmed_count": confirmed_count,
        "cancelled_count": cancelled_count,
        "conversion_rate": round(conversion_rate, 1),
        "recent_bookings": recent_bookings,
        "chart_tour_labels": chart_tour_labels,
        "chart_tour_data": chart_tour_data,
        "chart_monthly_labels": monthly_labels,
        "chart_monthly_data": monthly_sales,
    }
    return render(request, "admin/analytics.html", context)

@staff_member_required
def export_bookings_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passenger Bookings"

    # Header Styling
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid") # Amazon Green
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Booking Ref", "Customer Name", "Customer Email", "Status", 
        "Total Price", "Discount", "Grand Total", "Date Created"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add rows
    bookings = Booking.objects.all().order_by('-created_at')
    for b in bookings:
        traveler = b.travelers.first()
        cust_name = f"{traveler.first_name} {traveler.last_name}" if traveler else "Guest"
        cust_email = traveler.email if traveler else (b.user.email if b.user else "N/A")
        
        ws.append([
            str(b.booking_reference)[:8].upper(),
            cust_name,
            cust_email,
            b.get_status_display(),
            float(b.total_price),
            float(b.discount_amount),
            float(b.grand_total),
            b.created_at.strftime('%Y-%m-%d %H:%M') if b.created_at else ""
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=girasol_bookings_manifest.xlsx"
    wb.save(response)
    return response

@staff_member_required
def export_financial_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=girasol_financial_report.pdf"

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RepTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20, leading=24, textColor=colors.HexColor('#1B5E20')
    )
    body_style = ParagraphStyle(
        'RepBody', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, textColor=colors.HexColor('#263238')
    )
    header_style = ParagraphStyle(
        'RepHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=14, textColor=colors.white
    )

    story.append(Paragraph("GIRASOL FINANCIAL PERFORMANCE SUMMARY", title_style))
    story.append(Paragraph(f"Generated on: {datetime.today().strftime('%Y-%m-%d %H:%M')}", body_style))
    story.append(Spacer(1, 20))

    # Data Calculations
    total_sales = Booking.objects.filter(status='confirmed').aggregate(Sum('grand_total'))['grand_total__sum'] or 0.0
    total_discounts = Booking.objects.filter(status='confirmed').aggregate(Sum('discount_amount'))['discount_amount__sum'] or 0.0
    confirmed_count = Booking.objects.filter(status='confirmed').count()

    summary_data = [
        [Paragraph("<b>Confirmed Bookings Count:</b>", body_style), Paragraph(str(confirmed_count), body_style)],
        [Paragraph("<b>Total Promotions & Discounts:</b>", body_style), Paragraph(f"${total_discounts}", body_style)],
        [Paragraph("<b>Net Confirmed Revenue:</b>", body_style), Paragraph(f"${total_sales}", body_style)]
    ]
    summary_table = Table(summary_data, colWidths=[200, 300])
    summary_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CFD8DC')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 25))

    # List of bookings in table
    table_headers = [[
        Paragraph("Booking Ref", header_style),
        Paragraph("Status", header_style),
        Paragraph("Base Total", header_style),
        Paragraph("Discount", header_style),
        Paragraph("Grand Total", header_style),
        Paragraph("Date", header_style)
    ]]

    bookings = Booking.objects.all().order_by('-created_at')[:20] # Limit to recent 20 for brief summary report
    for b in bookings:
        table_headers.append([
            Paragraph(str(b.booking_reference)[:8].upper(), body_style),
            Paragraph(b.get_status_display(), body_style),
            Paragraph(f"${b.total_price}", body_style),
            Paragraph(f"${b.discount_amount}", body_style),
            Paragraph(f"${b.grand_total}", body_style),
            Paragraph(b.created_at.strftime('%Y-%m-%d'), body_style)
        ])

    bookings_table = Table(table_headers, colWidths=[100, 90, 80, 80, 80, 70])
    bookings_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E7D32')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#B0BEC5')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(Paragraph("<b>Recent Booking Breakdowns (Top 20)</b>", title_style))
    story.append(Spacer(1, 10))
    story.append(bookings_table)

    doc.build(story)
    return response
